# -*- coding: utf-8 -*-
# pip install requests beautifulsoup4 pandas openpyxl
import requests
from bs4 import BeautifulSoup
import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

URL = "https://www.mangazenkan.com/r/weekly/ebook/"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; mangazenkan-scraper/1.2)"
}


def fetch_weekly_ranking(url=URL):
    """漫画全巻ドットコムの週間ランキングを取得してDataFrameで返す"""
    res = requests.get(url, headers=HEADERS, timeout=20)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, "html.parser")

    items = soup.select("div.col-4.col-sm-4.col-md-3.col-lg-2")
    results = []

    for item in items:
        # --- ランク ---
        rank_elem = item.select_one("p.rank-number-small")
        rank = rank_elem.get_text(strip=True) if rank_elem else None

        # --- タイトル ---
        title_elem = item.select_one("div.product-name")
        title = title_elem.get_text(strip=True) if title_elem else None

        # --- 巻数（spanタグ構造も考慮）---
        volume = ""
        volume_elem = item.select_one("div.purchase-button-small")
        if volume_elem:
            # 例: "<span>4<small>巻</small></span>" にも対応
            vol_text = volume_elem.get_text(strip=True)
            m = re.search(r"(\d+)", vol_text)
            if m:
                volume = m.group(1)

        # --- 出版社（存在する場合）---
        publisher_elem = item.select_one("div.publisher")
        publisher = publisher_elem.get_text(strip=True) if publisher_elem else None

        if rank and title:
            results.append({
                "rank": int(rank),
                "title": title,
                "volume": int(volume) if volume.isdigit() else None,
                "publisher": publisher
            })

    return pd.DataFrame(results)


def save_to_excel(df, file_path="weekly_ranking.xlsx"):
    """取得結果をExcelに追記（シート名は実行日）、3巻以下を黄色でハイライト"""
    sheet_name = datetime.datetime.now().strftime("%Y-%m-%d")

    # --- Excel書き込み ---
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    except FileNotFoundError:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # --- 巻数3以下を黄色でハイライト ---
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # 色指定（淡い黄色）
    yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

    # "volume"列の列番号を特定
    volume_col = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value == "volume":
            volume_col = i
            break

    # 3巻以下をハイライト
    if volume_col:
        for row in ws.iter_rows(min_row=2, min_col=volume_col, max_col=volume_col):
            for cell in row:
                try:
                    if cell.value is not None and int(cell.value) <= 3:
                        cell.fill = yellow_fill
                except ValueError:
                    continue

    wb.save(file_path)
    print(f"✅ '{file_path}' にシート '{sheet_name}' を追加しました ({len(df)}件)")
    print("🎨 巻数が3以下の作品を黄色でハイライトしました！")


def main():
    df = fetch_weekly_ranking()
    save_to_excel(df)


if __name__ == "__main__":
    main()
